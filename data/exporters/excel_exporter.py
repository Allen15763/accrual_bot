"""
Excel匯出器

專門處理Excel格式的匯出功能，支援樣式設定和多工作表
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any, Union
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime

try:
    import xlsxwriter
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .base_exporter import BaseExporter
from ...core.models.config_models import ExportConfig
from ...utils.logging import Logger


@dataclass
class ExcelStyleConfig:
    """Excel樣式配置"""
    # 標題樣式
    header_font_bold: bool = True
    header_font_size: int = 12
    header_bg_color: str = 'D9D9D9'
    header_font_color: str = '000000'
    
    # 數據樣式
    data_font_size: int = 10
    data_font_color: str = '000000'
    
    # 邊框樣式
    border_style: str = 'thin'
    border_color: str = '000000'
    
    # 對齊方式
    header_alignment: str = 'center'
    data_alignment: str = 'left'
    
    # 數值格式
    number_format: str = '#,##0.00'
    date_format: str = 'yyyy-mm-dd'
    percentage_format: str = '0.00%'
    
    # 欄位寬度
    auto_adjust_width: bool = True
    min_column_width: int = 10
    max_column_width: int = 50


class ExcelExporter(BaseExporter):
    """Excel匯出器"""
    
    def __init__(self, config: Optional[ExportConfig] = None):
        super().__init__(config)
        
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel匯出需要安裝 openpyxl 和 xlsxwriter 套件")
        
        # Excel樣式配置
        self.style_config = ExcelStyleConfig()
        
        # 從ExportConfig更新樣式配置
        self._update_style_from_config()
    
    def _update_style_from_config(self):
        """從ExportConfig更新樣式配置"""
        if hasattr(self.config, 'header_style') and self.config.header_style:
            style = self.config.header_style
            if 'font_bold' in style:
                self.style_config.header_font_bold = style['font_bold']
            if 'bg_color' in style:
                self.style_config.header_bg_color = style['bg_color'].replace('#', '')
        
        if hasattr(self.config, 'data_style') and self.config.data_style:
            style = self.config.data_style
            if 'align' in style:
                self.style_config.data_alignment = style['align']
        
        if hasattr(self.config, 'number_format'):
            self.style_config.number_format = self.config.number_format
        
        if hasattr(self.config, 'date_format'):
            self.style_config.date_format = self.config.date_format
    
    def export(self, data: pd.DataFrame, output_path: Optional[str] = None) -> str:
        """
        匯出單一工作表的Excel檔案
        
        Args:
            data: 要匯出的DataFrame
            output_path: 輸出路徑
            
        Returns:
            str: 實際輸出路徑
        """
        if not self._validate_data(data):
            raise ValueError("數據驗證失敗")
        
        # 準備數據
        export_data = self._prepare_data_for_export(data)
        
        # 確定輸出路徑
        if output_path is None:
            filename = self._generate_filename(
                self.config.output_filename or "export", 
                "xlsx"
            )
            output_path = str(self.output_dir / filename)
        
        try:
            # 使用openpyxl進行高級格式化
            self._export_with_openpyxl(export_data, output_path)
            
            self.logger.info(f"成功匯出Excel檔案: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Excel匯出失敗: {e}")
            raise
    
    def _export_with_openpyxl(self, data: pd.DataFrame, output_path: str):
        """使用openpyxl匯出，支援進階格式化"""
        
        # 創建工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = self.config.excel_sheet_name
        
        # 寫入數據
        for r in dataframe_to_rows(data, index=self.config.include_index, header=True):
            worksheet.append(r)
        
        # 應用樣式
        self._apply_styles(worksheet, data)
        
        # 調整欄寬
        if self.style_config.auto_adjust_width:
            self._auto_adjust_column_width(worksheet, data)
        
        # 設定凍結窗格
        if self.config.freeze_panes:
            worksheet.freeze_panes = worksheet.cell(
                row=self.config.freeze_panes[0] + 1, 
                column=self.config.freeze_panes[1] + 1
            )
        
        # 設定篩選
        if self.config.enable_autofilter and len(data) > 0:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # 儲存檔案
        workbook.save(output_path)
    
    def _apply_styles(self, worksheet, data: pd.DataFrame):
        """應用樣式"""
        
        # 標題列樣式
        header_font = Font(
            bold=self.style_config.header_font_bold,
            size=self.style_config.header_font_size,
            color=self.style_config.header_font_color
        )
        
        header_fill = PatternFill(
            start_color=self.style_config.header_bg_color,
            end_color=self.style_config.header_bg_color,
            fill_type='solid'
        )
        
        header_alignment = Alignment(
            horizontal=self.style_config.header_alignment,
            vertical='center'
        )
        
        # 邊框樣式
        border = Border(
            left=Side(style=self.style_config.border_style, color=self.style_config.border_color),
            right=Side(style=self.style_config.border_style, color=self.style_config.border_color),
            top=Side(style=self.style_config.border_style, color=self.style_config.border_color),
            bottom=Side(style=self.style_config.border_style, color=self.style_config.border_color)
        )
        
        # 應用標題樣式
        for col in range(1, len(data.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 數據樣式
        data_font = Font(
            size=self.style_config.data_font_size,
            color=self.style_config.data_font_color
        )
        
        data_alignment = Alignment(
            horizontal=self.style_config.data_alignment,
            vertical='center'
        )
        
        # 應用數據樣式
        for row in range(2, len(data) + 2):
            for col in range(1, len(data.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # 根據數據類型設定格式
                column_name = data.columns[col - 1]
                self._apply_cell_format(cell, data, column_name, row - 2)
    
    def _apply_cell_format(self, cell, data: pd.DataFrame, column_name: str, row_index: int):
        """根據數據類型應用儲存格格式"""
        
        try:
            value = data.iloc[row_index, data.columns.get_loc(column_name)]
            
            # 數值格式
            if pd.api.types.is_numeric_dtype(data[column_name]):
                if any(keyword in column_name.lower() for keyword in ['amount', 'quantity', 'price', '金額', '數量', '價格']):
                    cell.number_format = self.style_config.number_format
            
            # 日期格式
            elif pd.api.types.is_datetime64_any_dtype(data[column_name]):
                cell.number_format = self.style_config.date_format
            
            # 百分比格式
            elif any(keyword in column_name.lower() for keyword in ['percent', 'rate', '%', '比例', '率']):
                cell.number_format = self.style_config.percentage_format
                
        except Exception as e:
            self.logger.warning(f"設定儲存格格式時發生錯誤 {column_name}: {e}")
    
    def _auto_adjust_column_width(self, worksheet, data: pd.DataFrame):
        """自動調整欄寬"""
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 設定欄寬，有最小和最大限制
            adjusted_width = min(
                max(max_length + 2, self.style_config.min_column_width),
                self.style_config.max_column_width
            )
            
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 使用配置中的特定欄寬設定
        if hasattr(self.config, 'column_widths') and self.config.column_widths:
            for col_name, width in self.config.column_widths.items():
                if col_name in data.columns:
                    col_index = data.columns.get_loc(col_name) + 1
                    col_letter = openpyxl.utils.get_column_letter(col_index)
                    worksheet.column_dimensions[col_letter].width = width
    
    def export_multiple_sheets(self, 
                             data_dict: Dict[str, pd.DataFrame], 
                             output_path: Optional[str] = None) -> str:
        """
        匯出多工作表Excel檔案
        
        Args:
            data_dict: {sheet_name: dataframe} 的字典
            output_path: 輸出路徑
            
        Returns:
            str: 輸出檔案路徑
        """
        if not data_dict:
            raise ValueError("沒有數據可匯出")
        
        # 確定輸出路徑
        if output_path is None:
            filename = self._generate_filename(
                self.config.output_filename or "export_multiple", 
                "xlsx"
            )
            output_path = str(self.output_dir / filename)
        
        try:
            # 創建工作簿
            workbook = openpyxl.Workbook()
            
            # 移除預設工作表
            workbook.remove(workbook.active)
            
            # 為每個數據創建工作表
            for sheet_name, data in data_dict.items():
                if not self._validate_data(data):
                    self.logger.warning(f"跳過無效的工作表: {sheet_name}")
                    continue
                
                # 準備數據
                export_data = self._prepare_data_for_export(data)
                
                # 創建工作表
                worksheet = workbook.create_sheet(title=sheet_name[:31])  # Excel工作表名稱限制31字符
                
                # 寫入數據
                for r in dataframe_to_rows(export_data, index=self.config.include_index, header=True):
                    worksheet.append(r)
                
                # 應用樣式
                self._apply_styles(worksheet, export_data)
                
                # 調整欄寬
                if self.style_config.auto_adjust_width:
                    self._auto_adjust_column_width(worksheet, export_data)
                
                # 設定凍結窗格
                if self.config.freeze_panes:
                    worksheet.freeze_panes = worksheet.cell(
                        row=self.config.freeze_panes[0] + 1, 
                        column=self.config.freeze_panes[1] + 1
                    )
                
                # 設定篩選
                if self.config.enable_autofilter and len(export_data) > 0:
                    worksheet.auto_filter.ref = worksheet.dimensions
            
            # 儲存檔案
            workbook.save(output_path)
            
            self.logger.info(f"成功匯出多工作表Excel檔案: {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"多工作表Excel匯出失敗: {e}")
            raise
    
    def supports_multiple_sheets(self) -> bool:
        """支援多工作表匯出"""
        return True
    
    def create_summary_sheet(self, data_dict: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """
        創建摘要工作表
        
        Args:
            data_dict: 數據字典
            
        Returns:
            pd.DataFrame: 摘要數據
        """
        summary_data = []
        
        for sheet_name, data in data_dict.items():
            summary_row = {
                '工作表名稱': sheet_name,
                '記錄數量': len(data),
                '欄位數量': len(data.columns),
                '建立時間': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # 添加數值欄位統計
            numeric_columns = data.select_dtypes(include=[np.number]).columns
            if len(numeric_columns) > 0:
                summary_row['數值欄位數'] = len(numeric_columns)
                summary_row['數值總和'] = data[numeric_columns].sum().sum()
            
            summary_data.append(summary_row)
        
        return pd.DataFrame(summary_data)


def export_to_excel(data: Union[pd.DataFrame, Dict[str, pd.DataFrame]], 
                   output_path: str,
                   config: Optional[ExportConfig] = None) -> str:
    """
    匯出到Excel的便捷函數
    
    Args:
        data: 要匯出的數據（單一DataFrame或多工作表字典）
        output_path: 輸出路徑
        config: 匯出配置
        
    Returns:
        str: 實際輸出路徑
    """
    exporter = ExcelExporter(config)
    
    if isinstance(data, dict):
        return exporter.export_multiple_sheets(data, output_path)
    else:
        return exporter.export(data, output_path)
