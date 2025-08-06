"""
Excel 專用數據導入器
提供Excel檔案的專業處理功能，包括複雜格式解析和數據預處理
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Union, Any
import openpyxl
from openpyxl import load_workbook

try:
    from .base_importer import BaseDataImporter
    from ...utils import get_logger, validate_file_path
except ImportError:
    # 如果相對導入失敗，使用絕對導入
    import sys
    from pathlib import Path
    
    # 添加accrual_bot目錄到sys.path
    current_dir = Path(__file__).parent.parent.parent
    if str(current_dir) not in sys.path:
        sys.path.insert(0, str(current_dir))  
    
    from data.importers.base_importer import BaseDataImporter
    from utils import get_logger, validate_file_path


class ExcelImporter(BaseDataImporter):
    """Excel專用數據導入器"""
    
    def __init__(self):
        """初始化Excel導入器"""
        super().__init__()
        self.logger = get_logger(self.__class__.__name__)
        
        self.logger.info("初始化Excel專用導入器")
    
    def import_excel_with_preprocessing(self, file_path: str, 
                                      preprocessing_config: Dict[str, Any] = None) -> Tuple[pd.DataFrame, int, int]:
        """
        導入Excel檔案並進行預處理，提取日期和月份信息
        
        Args:
            file_path: Excel檔案路徑
            preprocessing_config: 預處理配置
            
        Returns:
            Tuple[pd.DataFrame, int, int]: (數據, 日期YYYYMM, 月份)
        """
        try:
            if not validate_file_path(file_path):
                raise ValueError(f"無效的Excel檔案路徑: {file_path}")
            
            file_name = Path(file_path).name
            self.logger.info(f"開始導入並預處理Excel檔案: {file_name}")
            
            # 導入Excel數據
            df = self.import_file(file_path, **(preprocessing_config or {}))
            
            # 提取日期和月份信息
            date_info, month_info = self.extract_date_and_month_from_filename(file_name)
            
            if date_info is None or month_info is None:
                # 如果無法從檔案名提取，嘗試從數據中提取
                date_info, month_info = self._extract_date_from_data(df)
            
            # 應用預處理
            df = self._apply_preprocessing(df, preprocessing_config)
            
            self.logger.info(
                f"成功導入並預處理Excel檔案: {file_name}, "
                f"數據形狀: {df.shape}, 日期: {date_info}, 月份: {month_info}"
            )
            
            return df, date_info or 202501, month_info or 1
            
        except Exception as e:
            self.logger.error(f"導入並預處理Excel檔案失敗: {file_path}, 錯誤: {str(e)}", exc_info=True)
            raise
    
    def _extract_date_from_data(self, df: pd.DataFrame) -> Tuple[Optional[int], Optional[int]]:
        """從數據中提取日期信息"""
        try:
            # 尋找可能包含日期的欄位
            date_columns = [col for col in df.columns if any(keyword in str(col).lower() 
                          for keyword in ['date', '日期', 'month', '月份', 'period', '期間'])]
            
            for col in date_columns:
                if col in df.columns:
                    # 嘗試解析日期
                    date_series = pd.to_datetime(df[col], errors='coerce')
                    valid_dates = date_series.dropna()
                    
                    if not valid_dates.empty:
                        # 使用最常見的日期
                        most_common_date = valid_dates.mode()
                        if not most_common_date.empty:
                            date_obj = most_common_date.iloc[0]
                            year_month = int(date_obj.strftime('%Y%m'))
                            return year_month, date_obj.month
            
            return None, None
            
        except Exception as e:
            self.logger.warning(f"從數據中提取日期時出錯: {str(e)}")
            return None, None
    
    def _apply_preprocessing(self, df: pd.DataFrame, config: Dict[str, Any] = None) -> pd.DataFrame:
        """應用預處理邏輯"""
        try:
            if config is None:
                config = {}
            
            df_processed = df.copy()
            
            # 移除完全空白的行和列
            if config.get('remove_empty_rows', True):
                df_processed = df_processed.dropna(how='all')
            
            if config.get('remove_empty_columns', True):
                df_processed = df_processed.dropna(axis=1, how='all')
            
            # 重置索引
            if config.get('reset_index', True):
                df_processed = df_processed.reset_index(drop=True)
            
            # 清理欄位名稱
            if config.get('clean_column_names', True):
                df_processed.columns = [str(col).strip() for col in df_processed.columns]
            
            # 轉換數據類型
            if 'dtype_mapping' in config:
                for col, dtype in config['dtype_mapping'].items():
                    if col in df_processed.columns:
                        try:
                            df_processed[col] = df_processed[col].astype(dtype)
                        except Exception as e:
                            self.logger.warning(f"轉換欄位 {col} 的數據類型失敗: {str(e)}")
            
            # 填充空值
            if 'fill_na_values' in config:
                df_processed = df_processed.fillna(config['fill_na_values'])
            
            return df_processed
            
        except Exception as e:
            self.logger.error(f"應用預處理時出錯: {str(e)}", exc_info=True)
            return df
    
    def get_workbook_info(self, file_path: str) -> Dict[str, Any]:
        """
        獲取Excel工作簿的詳細信息
        
        Args:
            file_path: Excel檔案路徑
            
        Returns:
            Dict[str, Any]: 工作簿信息
        """
        try:
            if not validate_file_path(file_path):
                raise ValueError(f"無效的Excel檔案路徑: {file_path}")
            
            wb = load_workbook(file_path, read_only=True)
            
            info = {
                'filename': Path(file_path).name,
                'worksheets': [],
                'total_sheets': len(wb.worksheets)
            }
            
            for ws in wb.worksheets:
                sheet_info = {
                    'name': ws.title,
                    'max_row': ws.max_row,
                    'max_column': ws.max_column,
                    'has_data': ws.max_row > 1 or ws.max_column > 1
                }
                info['worksheets'].append(sheet_info)
            
            wb.close()
            return info
            
        except Exception as e:
            self.logger.error(f"獲取Excel工作簿信息時出錯: {str(e)}", exc_info=True)
            return {}
    
    def import_multiple_sheets(self, file_path: str, 
                              sheet_configs: Dict[str, Dict] = None) -> Dict[str, pd.DataFrame]:
        """
        從單個Excel檔案導入多個工作表
        
        Args:
            file_path: Excel檔案路徑
            sheet_configs: 每個工作表的配置
            
        Returns:
            Dict[str, pd.DataFrame]: 工作表名稱與DataFrame的對應字典
        """
        try:
            if not validate_file_path(file_path):
                raise ValueError(f"無效的Excel檔案路徑: {file_path}")
            
            if sheet_configs is None:
                sheet_configs = {}
            
            # 獲取工作簿信息
            workbook_info = self.get_workbook_info(file_path)
            sheet_names = [sheet['name'] for sheet in workbook_info.get('worksheets', [])]
            
            if not sheet_names:
                self.logger.warning(f"Excel檔案中沒有找到工作表: {file_path}")
                return {}
            
            results = {}
            
            self.logger.info(f"開始導入Excel檔案的 {len(sheet_names)} 個工作表")
            
            for sheet_name in sheet_names:
                try:
                    config = sheet_configs.get(sheet_name, {})
                    config['sheet_name'] = sheet_name
                    
                    df = self.import_file(file_path, **config)
                    
                    if not df.empty:
                        results[sheet_name] = df
                        
                except Exception as e:
                    self.logger.error(f"導入工作表失敗: {sheet_name}, 錯誤: {str(e)}")
                    continue
            
            self.logger.info(f"成功導入 {len(results)} 個工作表")
            return results
            
        except Exception as e:
            self.logger.error(f"導入多個工作表時出錯: {str(e)}", exc_info=True)
            return {}
    
    def import_with_data_validation(self, file_path: str, 
                                   validation_rules: Dict[str, Any] = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
        """
        導入Excel檔案並進行數據驗證
        
        Args:
            file_path: Excel檔案路徑
            validation_rules: 驗證規則
            
        Returns:
            Tuple[pd.DataFrame, Dict[str, Any]]: (數據, 驗證結果)
        """
        try:
            # 導入數據
            df = self.import_file(file_path)
            
            # 進行數據驗證
            validation_result = self._validate_data(df, validation_rules or {})
            
            return df, validation_result
            
        except Exception as e:
            self.logger.error(f"導入並驗證Excel檔案時出錯: {str(e)}", exc_info=True)
            raise
    
    def _validate_data(self, df: pd.DataFrame, rules: Dict[str, Any]) -> Dict[str, Any]:
        """驗證數據"""
        try:
            result = {
                'is_valid': True,
                'errors': [],
                'warnings': [],
                'statistics': {
                    'total_rows': len(df),
                    'total_columns': len(df.columns),
                    'empty_rows': df.isnull().all(axis=1).sum(),
                    'empty_columns': df.isnull().all(axis=0).sum()
                }
            }
            
            # 檢查必要欄位
            required_columns = rules.get('required_columns', [])
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                result['errors'].append(f"缺少必要欄位: {missing_columns}")
                result['is_valid'] = False
            
            # 檢查最少行數
            min_rows = rules.get('min_rows', 0)
            if len(df) < min_rows:
                result['errors'].append(f"數據行數不足: {len(df)} < {min_rows}")
                result['is_valid'] = False
            
            # 檢查數據類型
            if 'column_types' in rules:
                for col, expected_type in rules['column_types'].items():
                    if col in df.columns:
                        try:
                            df[col].astype(expected_type)
                        except Exception:
                            result['warnings'].append(f"欄位 {col} 的數據類型可能不正確")
            
            # 檢查空值
            if 'null_check_columns' in rules:
                for col in rules['null_check_columns']:
                    if col in df.columns:
                        null_count = df[col].isnull().sum()
                        if null_count > 0:
                            result['warnings'].append(f"欄位 {col} 包含 {null_count} 個空值")
            
            return result
            
        except Exception as e:
            self.logger.error(f"驗證數據時出錯: {str(e)}")
            return {'is_valid': False, 'errors': [str(e)], 'warnings': [], 'statistics': {}}
    
    def concurrent_import_excel_files(self, file_paths: List[str], 
                                     preprocessing_configs: Dict[str, Dict] = None) -> Dict[str, Tuple[pd.DataFrame, int, int]]:
        """
        並發導入多個Excel檔案並進行預處理
        
        Args:
            file_paths: Excel檔案路徑列表
            preprocessing_configs: 每個檔案的預處理配置
            
        Returns:
            Dict[str, Tuple[pd.DataFrame, int, int]]: 檔案名稱與(數據, 日期, 月份)的對應字典
        """
        try:
            if not file_paths:
                return {}
            
            if preprocessing_configs is None:
                preprocessing_configs = {}
            
            # 過濾有效的檔案路徑
            valid_files = [(path, Path(path).stem) for path in file_paths if path and validate_file_path(path)]
            
            if not valid_files:
                self.logger.warning("沒有有效的Excel檔案路徑")
                return {}
            
            self.logger.info(f"開始並發導入 {len(valid_files)} 個Excel檔案")
            
            results = {}
            
            # 使用線程池並發處理
            import concurrent.futures
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                # 提交任務
                future_to_file = {}
                for file_path, file_name in valid_files:
                    config = preprocessing_configs.get(file_name, {})
                    future = executor.submit(self._import_excel_safe, file_path, file_name, config)
                    future_to_file[future] = file_name
                
                # 收集結果
                for future in concurrent.futures.as_completed(future_to_file, timeout=self.timeout):
                    file_name = future_to_file[future]
                    try:
                        result = future.result()
                        if result is not None:
                            results[file_name] = result
                    except Exception as e:
                        self.logger.error(f"併發導入Excel檔案失敗: {file_name}, 錯誤: {str(e)}")
            
            self.logger.info(f"並發導入完成，成功導入 {len(results)} 個Excel檔案")
            return results
            
        except Exception as e:
            self.logger.error(f"並發導入Excel檔案時出錯: {str(e)}", exc_info=True)
            return {}
    
    def _import_excel_safe(self, file_path: str, file_name: str, 
                          config: Dict) -> Optional[Tuple[pd.DataFrame, int, int]]:
        """安全地導入Excel檔案（用於並發處理）"""
        try:
            return self.import_excel_with_preprocessing(file_path, config)
        except Exception as e:
            self.logger.error(f"導入Excel檔案失敗: {file_name}, 錯誤: {str(e)}")
            return None
